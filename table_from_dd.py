import mysql.connector
import openpyxl
from mysql.connector import errorcode
from python_mysql_dbconfig import read_db_config
from pathlib import Path

dBConnector = mysql.connector.connect(**read_db_config())
dBCursor = dBConnector.cursor()

dataDictionary_file = Path("", "Data_Dictionary_software_control.xlsx")
dd_obj = openpyxl.load_workbook(dataDictionary_file)
sheet = dd_obj['tutorial_tble1']

dd_parse = {}
r = 0
for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=True):
    c = 0
    if r == 0:
        dd_header = row
    if r > 0:
        dd_parse.setdefault(row[0], {})
        for value in row:
            dd_parse[row[0]][dd_header[c]] = value
            c += 1
    r += 1

initTables = {}
sqlCommand = []
sqlCommand.append("CREATE TABLE tutorial_tble1(")
for key in dd_parse:
    sqlCommand.append(key)
    sqlCommand.append(dd_parse[key]['Data Type'])
    if dd_parse[key]['Auto_Increment'] == "T":
        sqlCommand.append("AUTO_INCREMENT")
    if dd_parse[key]['NULL'] == "F":
        sqlCommand.append("NOT NULL")
    sqlCommand.append(",")

match = False
for key in dd_parse:
    if dd_parse[key]['Primary Key'] == "m":
        match = True
        sqlCommand.append("PRIMARY KEY")
        sqlCommand.append("(")
        sqlCommand.append(dd_parse[key]['Attribute'])
        sqlCommand.append(")")
if not match:
    sqlCommand.pop()
sqlCommand.append(") ENGINE=InnoDB")

initTables['tutorial_tble1'] = ' '.join(sqlCommand)
dBName = "test_db1"

# initTables = {}
# initTables['tutorial_tble1'] = (
#     "CREATE TABLE tutorial_tble1("
#     "    tutorial_id INT NOT NULL AUTO_INCREMENT,"
#     "    tutorial_title VARCHAR(100) NOT NULL,"
#     "    tutorial_author VARCHAR(40) NOT NULL,"
#     "    submission_date DATE,"
#     "    PRIMARY KEY (tutorial_id)"
#     ") ENGINE=InnoDB")

for table_name in initTables:
    table_description = initTables[table_name]
    try:
        #print("Creating table {}: ".format(table_name), end='')
        print(table_description)
        dBCursor.execute(table_description)
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_TABLE_EXISTS_ERROR:
            print("already exists.")
        else:
            print(err.msg)
    else:
        print("OK")

dBCursor.close()
dBConnector.close()